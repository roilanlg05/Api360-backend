from __future__ import annotations

import asyncio
import logging
import re
from shared.models.trip import Trip
from datetime import datetime, date, time, timedelta
from typing import Any, BinaryIO, List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# --- Regex para vuelos ---

# Vuelos tipo: WN 2453, AA1234, WN 2453-01, etc.
FLIGHT_ONLY_RE = re.compile(
    r"(?P<flight>[A-Z0-9]{2}\s*\d{3,4})(?:-\d{2})?"
)

# Vuelos + fecha opcional + hora:
# - "WN 2668-01 Nov 04:55"
# - "WN 4285-16 Nov 04:45"
# - también soporta sin mes: "WN 1322-01 17:20"
FLIGHT_TIME_RE = re.compile(
    r"(?P<flight>[A-Z0-9]{2}\s*\d{3,4}(?:-\d{2})?)\s+(?:[A-Za-z]{3}\s+)?(?P<hour>\d{1,2}):(?P<minute>\d{2})"
)


# ---------- Helpers básicos ----------

def _normalize_str(value: Any) -> str:
    """Normaliza un valor a string limpio."""
    if value is None:
        return ""
    return str(value).strip()


def _find_city_code(ws) -> Optional[str]:
    """
    Busca en las primeras filas algo como:
    CITY:    SDF

    Devuelve el código de ciudad/airport (ej. 'SDF') o None si no lo encuentra.
    """
    max_rows_to_scan = 40

    for row in ws.iter_rows(min_row=1, max_row=max_rows_to_scan):
        for cell in row:
            text = _normalize_str(cell.value).upper()
            if not text:
                continue

            if text in ("CITY", "CITY:"):
                row_idx = cell.row
                col_idx = cell.column

                # Buscar hacia la derecha
                for other_col in range(col_idx + 1, col_idx + 20):
                    other_cell = ws.cell(row=row_idx, column=other_col)
                    val = _normalize_str(other_cell.value)
                    if val:
                        return val.upper()

                # Si no encontramos nada a la derecha, probar otras celdas
                values_in_row = [
                    _normalize_str(c.value)
                    for c in row
                    if _normalize_str(c.value) and _normalize_str(c.value).upper() not in ("CITY", "CITY:")
                ]
                if values_in_row:
                    return values_in_row[0].upper()

    return None


def _parse_service_date(value: Any) -> date:
    """
    Convierte la columna de fecha a date.
    Soporta date/datetime o strings tipo:
    "01-Nov-2025", "2025-11-01", "11/01/2025", etc.
    """
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    text = _normalize_str(value)
    if not text:
        raise ValueError("Fecha de servicio vacía")

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%d-%B-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError(f"Formato de fecha no reconocido: {text!r}")


def _parse_flight_only(value: str) -> Optional[str]:
    """
    Extrae vuelo tipo WN 2453-01 de un texto.
    Si no hay vuelo, devuelve None.
    """
    text = _normalize_str(value)
    if not text:
        return None

    m = FLIGHT_ONLY_RE.search(text)
    if not m:
        return None

    return m.group("flight")


def _parse_flight_and_time(value: str, service_date: date) -> Tuple[Optional[str], Optional[datetime]]:
    """
    Extrae vuelo + hora de un texto tipo:
    - 'WN 2668-01 Nov 04:55'
    - 'WN 4285-16 Nov 04:45'
    - 'WN 1322-01 17:20'

    Si no encuentra patrón, devuelve (None, None).
    """
    text = _normalize_str(value)
    if not text:
        return None, None

    m = FLIGHT_TIME_RE.search(text)
    if not m:
        return None, None

    flight = m.group("flight")
    hour = int(m.group("hour"))
    minute = int(m.group("minute"))

    dt = datetime(
        year=service_date.year,
        month=service_date.month,
        day=service_date.day,
        hour=hour,
        minute=minute,
    )

    return flight, dt


def _split_airline_and_number(flight: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """
    Separa 'WN 2453-01' en:
    - airline: 'WN'
    - flight_number: '2453'
    """
    if not flight:
        return None, None

    text = _normalize_str(flight)
    m = re.match(r"^(?P<airline>[A-Z0-9]{2})\s*(?P<number>\d{1,4})(?:-\d{2})?", text)
    if not m:
        return None, text

    return m.group("airline"), m.group("number")


# ---------- Detección de encabezados ----------

def _find_header_and_subheader(ws) -> Tuple[int, Any, Any]:
    """
    Busca la fila donde aparecen 'DATE', 'PICK UP', 'DROP OFF'
    y devuelve (row_index, header_row, subheader_row).
    """
    for row in ws.iter_rows(min_row=1, max_row=80):
        texts = [_normalize_str(c.value).upper() for c in row if c.value is not None]
        if not texts:
            continue

        has_date = "DATE" in texts
        has_pickup = any("PICK" in t and "UP" in t for t in texts)
        has_dropoff = any("DROP" in t and "OFF" in t for t in texts)

        if has_date and has_pickup and has_dropoff:
            header_row_index = row[0].row
            header_row = row
            subheader_row = ws[header_row_index + 1]
            return header_row_index, header_row, subheader_row

    raise RuntimeError(
        "No se encontró la fila de encabezados (con DATE / PICK UP / DROP OFF) en la hoja Schedule."
    )


def _determine_columns(header_row, subheader_row) -> dict:
    """
    A partir de encabezado y subencabezado, determina las columnas.
    """
    date_col = None
    riders_col = None

    for cell in header_row:
        text = _normalize_str(cell.value).upper()
        if text == "DATE":
            date_col = cell.column
        if "RIDERS" in text:
            riders_col = cell.column

    if date_col is None:
        raise RuntimeError("No se encontró la columna DATE en el encabezado.")

    pickup_loc_col = None
    dropoff_loc_col = None
    from_col = None
    to_col = None

    for cell in subheader_row:
        text = _normalize_str(cell.value).upper()
        col = cell.column

        if text == "LOCATION":
            if pickup_loc_col is None:
                pickup_loc_col = col
            else:
                dropoff_loc_col = col
        elif "FROM" in text:
            from_col = col
        elif text == "TO":
            to_col = col

    if from_col is None or to_col is None:
        raise RuntimeError("No se encontraron las columnas 'From' y 'To' en el subencabezado.")

    return {
        "date": date_col,
        "riders": riders_col,
        "pickup_from": from_col,
        "dropoff_to": to_col,
        "pickup_location": pickup_loc_col,
        "dropoff_location": dropoff_loc_col,
    }


# ---------- Procesamiento sincrónico (ejecutado en thread) ----------

def _process_excel_sync(stream: BinaryIO, sheet_name: str) -> Tuple[List[Trip], Optional[str]]:
    """
    Función sincrónica que procesa el Excel.
    Se ejecutará en un thread pool para no bloquear el event loop.
    """
    wb = load_workbook(stream, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"No se encontró la hoja {sheet_name!r} en el archivo de Excel.")

    ws = wb[sheet_name]

    city_code = _find_city_code(ws)
    header_row_index, header_row, subheader_row = _find_header_and_subheader(ws)
    cols = _determine_columns(header_row, subheader_row)

    trips: List[Trip] = []
    current_service_date: Optional[date] = None

    for row in ws.iter_rows(min_row=header_row_index + 2):
        row_idx = row[0].row

        def get(col_name: str):
            col_idx = cols.get(col_name)
            if not col_idx:
                return None
            return ws.cell(row=row_idx, column=col_idx).value

        # --- DATE con arrastre ---
        date_raw = get("date")
        date_str = _normalize_str(date_raw)

        upper_date_str = date_str.upper()
        if upper_date_str and (
            "END OF TRANSPORTATION SCHEDULE" in upper_date_str
            or "THIS SCHEDULE REPRESENTS" in upper_date_str
        ):
            break

        if date_str:
            try:
                service_date = _parse_service_date(date_raw)
                current_service_date = service_date
            except Exception as exc:
                logger.error("Error parseando fecha en fila %s: %s", row_idx, exc)
                continue
        else:
            if current_service_date is None:
                continue
            service_date = current_service_date

        # --- Otras columnas ---
        pickup_val = get("pickup_from")
        dropoff_val = get("dropoff_to")
        pickup_loc_val = get("pickup_location")
        dropoff_loc_val = get("dropoff_location")
        riders_val = get("riders")

        pickup_raw = _normalize_str(pickup_val)
        dropoff_raw = _normalize_str(dropoff_val)

        if not any([pickup_raw, dropoff_raw, pickup_loc_val, dropoff_loc_val]):
            continue

        try:
            # Determinar código de aeropuerto
            airport_code: Optional[str] = None
            for v in (pickup_loc_val, dropoff_loc_val):
                s = _normalize_str(v).upper()
                if len(s) == 3 and s.isalpha():
                    airport_code = s
                    break

            # Riders
            riders = 0
            if riders_val not in (None, ""):
                try:
                    riders = int(float(str(riders_val).strip()))
                except ValueError:
                    logger.warning(
                        "No se pudo convertir riders=%r a int en fila %s, se usará 0",
                        riders_val,
                        row_idx,
                    )
                    riders = 0

            # ¿Pick Up (From) es un vuelo?
            flight_from_pickup = _parse_flight_only(pickup_raw)

            if flight_from_pickup:
                # ----- Caso: PICK UP EN AEROPUERTO -----
                airline, flight_number = _split_airline_and_number(flight_from_pickup)

                pick_up_location = airport_code or "AIRPORT"

                dropoff_flight = _parse_flight_only(dropoff_raw)
                if dropoff_flight and airport_code:
                    drop_off_location = airport_code
                else:
                    drop_off_location = dropoff_raw or (airport_code or "AIRPORT")

                _, dropoff_dt = _parse_flight_and_time(dropoff_raw, service_date)
                if dropoff_dt is None:
                    _, pickup_dt = _parse_flight_and_time(pickup_raw, service_date)
                else:
                    pickup_dt = dropoff_dt

                if pickup_dt is None:
                    pickup_dt = datetime(
                        year=service_date.year,
                        month=service_date.month,
                        day=service_date.day,
                        hour=0,
                        minute=0,
                    )

                pick_up_date = pickup_dt.date()
                pick_up_time = pickup_dt.time()

            else:
                # ----- Caso: PICK UP EN HOTEL -----
                pick_up_location = pickup_raw

                flight_from_dropoff, dropoff_dt = _parse_flight_and_time(dropoff_raw, service_date)
                if not flight_from_dropoff or not dropoff_dt:
                    raise ValueError(f"Formato inválido de vuelo en Drop Off: {dropoff_raw!r}")

                airline, flight_number = _split_airline_and_number(flight_from_dropoff)

                pickup_dt = dropoff_dt - timedelta(minutes=20)
                pick_up_date = pickup_dt.date()
                pick_up_time = pickup_dt.time()

                drop_off_location = airport_code or "AIRPORT"

            trips.append(
                Trip(
                    pick_up_date=pick_up_date,
                    pick_up_time=pick_up_time,
                    pick_up_location=pick_up_location,
                    drop_off_location=drop_off_location,
                    airline=airline,
                    flight_number=flight_number,
                    riders=riders
                )
            )

        except Exception as exc:
            logger.error("Error parseando fila en hoja 'Schedule': %s", exc)
            continue

    wb.close()
    return trips, city_code


# ---------- API pública asíncrona ----------

async def load_trips_from_excel(
    stream: BinaryIO, 
    sheet_name: str = "Schedule"
) -> tuple[list[Trip], Optional[str]]:
    """
    Lee un Excel estilo Air Crew Transport / SDF y devuelve una lista de Trip.
    
    Esta función es asíncrona y ejecuta el procesamiento del Excel en un 
    thread pool para no bloquear el event loop.

    Args:
        stream: Archivo Excel como BinaryIO
        sheet_name: Nombre de la hoja a procesar (default: "Schedule")

    Returns:
        Tuple[List[Trip], Optional[str]]: Lista de trips y código de ciudad
    
    Raises:
        RuntimeError: Si no se encuentra la hoja o los encabezados esperados
    """
    return await asyncio.to_thread(_process_excel_sync, stream, sheet_name)


async def load_trips_from_bytes(
    data: bytes, 
    sheet_name: str = "Schedule"
) -> tuple[list[Trip], Optional[str]]:
    """
    Versión que acepta bytes directamente (útil para FastAPI UploadFile).
    
    Args:
        data: Contenido del archivo Excel como bytes
        sheet_name: Nombre de la hoja a procesar

    Returns:
        Tuple[List[Trip], Optional[str]]: Lista de trips y código de ciudad
    """
    from io import BytesIO
    stream = BytesIO(data)
    return await load_trips_from_excel(stream, sheet_name)
